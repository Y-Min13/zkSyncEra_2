import openpyxl
import os.path
import settings as stgs
from loguru import logger as cs_logger
from sys import stderr

cs_logger.remove()
cs_logger.add(stderr, format="<white>{time:HH:mm:ss}</white> | <level>{level: <8}</level> | <white>{message}</white>")


def create_xml(log_file):
    if os.path.exists(log_file) is False:
        workbook = openpyxl.Workbook()
        workbook.save(log_file)

        worksheet = workbook.active
        worksheet.title = 'Overall'
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "Адрес кошелька"
        worksheet.cell(row=1, column=3).value = "Адрес биржи"
        worksheet.cell(row=1, column=4).value = "Сумма бриджа"
        worksheet.cell(row=1, column=5).value = "Сумма свапов"
        worksheet.cell(row=1, column=6).value = "Баланс до zkSyncEra"
        worksheet.cell(row=1, column=7).value = "Баланс после zkSyncEra"
        worksheet.cell(row=1, column=8).value = "Кол-во транз в скрипте"
        worksheet.cell(row=1, column=9).value = "Кол-во транз кошелька"
        worksheet.cell(row=1, column=10).value = "Нач. баланс биржи"
        worksheet.cell(row=1, column=11).value = "Кон. баланс биржи"
        worksheet.cell(row=1, column=12).value = "Время"
        workbook.save(log_file)

        workbook.create_sheet('Swap operations')
        worksheet = workbook['Swap operations']
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "№ операции"
        worksheet.cell(row=1, column=3).value = "Модуль"
        worksheet.cell(row=1, column=4).value = "Адрес кошелька"
        worksheet.cell(row=1, column=5).value = "MuteSwap txns"
        worksheet.cell(row=1, column=6).value = "SyncSwap txns"
        worksheet.cell(row=1, column=7).value = "SpacefiSwap txns"
        worksheet.cell(row=1, column=8).value = "PancakeSwap txns"
        worksheet.cell(row=1, column=9).value = "MaverickSwap txns"
        worksheet.cell(row=1, column=10).value = "MuteSwap ETH сумма"
        worksheet.cell(row=1, column=11).value = "SyncSwap ETH сумма"
        worksheet.cell(row=1, column=12).value = "SpacefiSwap ETH сумма"
        worksheet.cell(row=1, column=13).value = "PancakeSwap ETH сумма"
        worksheet.cell(row=1, column=14).value = "MaverickSwap ETH сумма"
        worksheet.cell(row=1, column=15).value = "Начальный баланс ETH"
        worksheet.cell(row=1, column=16).value = "Конечный баланс ETH"
        worksheet.cell(row=1, column=17).value = "Начальный баланс USDC"
        worksheet.cell(row=1, column=18).value = "Конечный баланс USDC"
        worksheet.cell(row=1, column=19).value = "Время"
        workbook.save(log_file)

        worksheet = workbook.create_sheet('Swap transactions')
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "№ операции"
        worksheet.cell(row=1, column=3).value = "№ транзакции"
        worksheet.cell(row=1, column=4).value = "Адрес кошелька"
        worksheet.cell(row=1, column=5).value = "Свап"
        worksheet.cell(row=1, column=6).value = "Сумма свапа ETH"
        worksheet.cell(row=1, column=7).value = "Сумма свапа USDC"
        worksheet.cell(row=1, column=8).value = "Hash свапа"
        worksheet.cell(row=1, column=9).value = "Начальный баланс ETH"
        worksheet.cell(row=1, column=10).value = "Конечный баланс ETH"
        worksheet.cell(row=1, column=11).value = "Начальный баланс USDC"
        worksheet.cell(row=1, column=12).value = "Конечный баланс USDC"
        worksheet.cell(row=1, column=13).value = "Время"
        workbook.save(log_file)

        worksheet = workbook.create_sheet('Bridge transactions')
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "Адрес"
        worksheet.cell(row=1, column=3).value = "Исх. сеть"
        worksheet.cell(row=1, column=4).value = "Вхд. сеть"
        worksheet.cell(row=1, column=5).value = f"Hash бриджа"
        worksheet.cell(row=1, column=6).value = f"Начальный баланс исх. сети"
        worksheet.cell(row=1, column=7).value = f"Конечный баланс исх. сети"
        worksheet.cell(row=1, column=8).value = "Сумма бриджа ETH"
        worksheet.cell(row=1, column=9).value = f"Начальный баланс вхд. сети"
        worksheet.cell(row=1, column=10).value = f"Конечный баланс вхд. сети"
        worksheet.cell(row=1, column=11).value = "Время"
        workbook.save(log_file)
        workbook.close()


def write_balance_sm(log_file, wallet_num, operation, module, address, mute_txns, mute_value, sync_txns, sync_value,
                     spacefi_txns, spacefi_value, pancake_txns, pancake_value,  maverick_txns, maverick_value,
                     balance_start_eth, balance_end_eth, balance_start_token, balance_end_token, script_time):
    while True:
        try:
            workbook = openpyxl.load_workbook(log_file)
            worksheet = workbook['Swap operations']
            last_row = worksheet.max_row
            worksheet.cell(row=last_row + 1, column=1).value = wallet_num
            worksheet.cell(row=last_row + 1, column=2).value = operation
            worksheet.cell(row=last_row + 1, column=3).value = module
            worksheet.cell(row=last_row + 1, column=4).value = address
            worksheet.cell(row=last_row + 1, column=5).value = mute_txns
            worksheet.cell(row=last_row + 1, column=6).value = sync_txns
            worksheet.cell(row=last_row + 1, column=7).value = spacefi_txns
            worksheet.cell(row=last_row + 1, column=8).value = pancake_txns
            worksheet.cell(row=last_row + 1, column=9).value = maverick_txns
            worksheet.cell(row=last_row + 1, column=10).value = mute_value
            worksheet.cell(row=last_row + 1, column=10).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=11).value = sync_value
            worksheet.cell(row=last_row + 1, column=11).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=12).value = spacefi_value
            worksheet.cell(row=last_row + 1, column=12).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=13).value = pancake_value
            worksheet.cell(row=last_row + 1, column=13).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=14).value = maverick_value
            worksheet.cell(row=last_row + 1, column=14).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=15).value = balance_start_eth
            worksheet.cell(row=last_row + 1, column=15).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=16).value = balance_end_eth
            worksheet.cell(row=last_row + 1, column=16).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=17).value = balance_start_token
            worksheet.cell(row=last_row + 1, column=17).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=18).value = balance_end_token
            worksheet.cell(row=last_row + 1, column=18).number_format = '0.00000'
            worksheet.cell(row=last_row + 1, column=19).value = script_time
            workbook.save(log_file)
            workbook.close()
            break
        except PermissionError:
            cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
            input("")


def get_last_row_sm(log_file):
    workbook = openpyxl.load_workbook(log_file)
    worksheet = workbook['Swap operations']
    last_row = worksheet.max_row
    if last_row > 1:
        index = worksheet.cell(row=last_row, column=2).value + 1
    else:
        index = 1
    return index


def get_last_row_overall(log_file):
    workbook = openpyxl.load_workbook(log_file)
    worksheet = workbook['Overall']
    last_row = worksheet.max_row
    #if last_row > 1:
    #    index = worksheet.cell(row=last_row, column=1).value
    #else:
    #    index = 1
    return last_row


def write_overall(wallet, wallet_num, bridge_value, swap_value, balance_st, balance_end, script_time, nonce):
    while True:
        try:
            workbook = openpyxl.load_workbook(stgs.log_file)
            worksheet = workbook['Overall']
            last_row = stgs.last_row
            worksheet.cell(row=last_row + wallet.index, column=1).value = f'{wallet_num}'
            worksheet.cell(row=last_row + wallet.index, column=2).value = wallet.address
            worksheet.cell(row=last_row + wallet.index, column=3).value = wallet.exchange_address

            worksheet.cell(row=last_row + wallet.index, column=4).value = bridge_value
            worksheet.cell(row=last_row + wallet.index, column=4).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=5).value = swap_value
            worksheet.cell(row=last_row + wallet.index, column=5).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=6).value = balance_st
            worksheet.cell(row=last_row + wallet.index, column=6).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=7).value = balance_end
            worksheet.cell(row=last_row + wallet.index, column=7).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=8).value = wallet.txn_num
            worksheet.cell(row=last_row + wallet.index, column=9).value = nonce

            worksheet.cell(row=last_row + wallet.index, column=10).value = wallet.exc_bal_st
            worksheet.cell(row=last_row + wallet.index, column=10).number_format = '0.00000'

            worksheet.cell(row=last_row + wallet.index, column=11).value = wallet.exc_bal_end
            worksheet.cell(row=last_row + wallet.index, column=11).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=12).value = script_time
            workbook.save(stgs.log_file)
            workbook.close()
            break
        except PermissionError:
            cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
            input("")


def rewrite_overall(wallet, bridge_value, swap_value, balance_end, nonce):
    while True:
        try:
            workbook = openpyxl.load_workbook(stgs.log_file)
            worksheet = workbook['Overall']
            last_row = stgs.last_row
            worksheet.cell(row=last_row + wallet.index, column=4).value = bridge_value
            worksheet.cell(row=last_row + wallet.index, column=5).value = swap_value
            worksheet.cell(row=last_row + wallet.index, column=7).value = balance_end
            worksheet.cell(row=last_row + wallet.index, column=8).value = wallet.txn_num
            worksheet.cell(row=last_row + wallet.index, column=9).value = nonce
            worksheet.cell(row=last_row + wallet.index, column=10).value = wallet.exc_bal_st
            worksheet.cell(row=last_row + wallet.index, column=11).value = wallet.exc_bal_end
            workbook.save(stgs.log_file)
            workbook.close()
            break
        except PermissionError:
            cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
            input("")


class LogSwap(object):
    def __init__(self, wallet_num, operation, txn_num, address, swapper, swap_value, hash_txn, balance_start_eth,
                 balance_end_eth, balance_start_token, balance_end_token):
        self.wallet_num = wallet_num
        self.operation = operation
        self.txn_num = txn_num
        self.address = address
        self.swapper = swapper
        self.swap_value = swap_value
        self.hash_txn = hash_txn
        self.balance_start_eth = balance_start_eth
        self.balance_end_eth = balance_end_eth
        self.balance_start_token = balance_start_token
        self.balance_end_token = balance_end_token

    def write_log(self, log_file, token, script_time):
        while True:
            try:
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Swap transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row + 1, column=1).value = self.wallet_num
                worksheet.cell(row=last_row + 1, column=2).value = self.operation
                worksheet.cell(row=last_row + 1, column=3).value = self.txn_num
                worksheet.cell(row=last_row + 1, column=4).value = f'{self.address}'
                worksheet.cell(row=last_row + 1, column=5).value = f'{self.swapper}'

                if token == 1:
                    worksheet.cell(row=last_row + 1, column=6).value = self.swap_value
                    worksheet.cell(row=last_row + 1, column=6).number_format = '0.00000'
                    worksheet.cell(row=last_row + 1, column=7).value = ''
                if token == 2:
                    worksheet.cell(row=last_row + 1, column=6).value = ''
                    worksheet.cell(row=last_row + 1, column=7).value = self.swap_value
                    worksheet.cell(row=last_row + 1, column=7).number_format = '0.00000'

                worksheet.cell(row=last_row + 1, column=8).value = f'{self.hash_txn}'
                worksheet.cell(row=last_row + 1, column=9).value = self.balance_start_eth
                worksheet.cell(row=last_row + 1, column=9).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=10).value = self.balance_end_eth
                worksheet.cell(row=last_row + 1, column=10).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=11).value = self.balance_start_token
                worksheet.cell(row=last_row + 1, column=11).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=12).value = self.balance_end_token
                worksheet.cell(row=last_row + 1, column=12).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=13).value = f'{script_time}'
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")


class LogBridge(object):
    def __init__(self, index, net_from, net_to, address, bridge_value,
                 txn_hash_from, balance_from_st, balance_to_st, balance_from_end, balance_to_end):
        self.address = address
        self.bridge_value = bridge_value
        self.txn_hash_from = txn_hash_from
        self.balance_from_st = balance_from_st
        self.balance_to_st = balance_to_st
        self.balance_from_end = balance_from_end
        self.balance_to_end = balance_to_end
        self.index = index
        self.net_from = net_from[0]
        self.net_to = net_to[0]

    def write_log(self, log_file, script_time):
        while True:
            try:
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Bridge transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row + 1, column=1).value = self.index
                worksheet.cell(row=last_row + 1, column=2).value = f'{self.address}'
                worksheet.cell(row=last_row + 1, column=3).value = f'{self.net_from}'
                worksheet.cell(row=last_row + 1, column=4).value = f'{self.net_to}'
                worksheet.cell(row=last_row + 1, column=5).value = f'{self.txn_hash_from}'
                worksheet.cell(row=last_row + 1, column=6).value = self.balance_from_st
                worksheet.cell(row=last_row + 1, column=6).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=7).value = self.balance_from_end
                worksheet.cell(row=last_row + 1, column=7).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=8).value = self.bridge_value
                worksheet.cell(row=last_row + 1, column=8).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=9).value = self.balance_to_st
                worksheet.cell(row=last_row + 1, column=9).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=10).value = self.balance_to_end
                worksheet.cell(row=last_row + 1, column=10).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=11).value = f'{script_time}'
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")

    def rewrite_log(self, log_file):
        while True:
            try:
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Bridge transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row, column=7).value = self.balance_from_end
                worksheet.cell(row=last_row, column=10).value = self.balance_to_end
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")
